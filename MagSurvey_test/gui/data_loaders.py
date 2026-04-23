import threading
from pathlib import Path
import pandas as pd
from tkinter import messagebox

from logic.survey_processor import process_survey_folder
from logic.navigation_processor import process_navigation_folder
from logic.excel_writer import save_survey_excels
from logic.coordinate_merger import parse_navigation_text

from gui.map_manager import MapManager 

class DataLoaders:
    """Методы для загрузки данных в фоновом потоке и обновления UI."""
    
    def __init__(self, main_window):
        self.mw = main_window   # ссылка на MainWindow для доступа к переменным и методам
    
    def load_survey(self):
        if not self.mw.survey_path.get():
            return
        if self.mw.mode.get() == "text":
            messagebox.showinfo("Информация", "Режим 'Текстовый файл' находится в разработке")
            return
        
        self.mw.status_var.set("Загрузка данных съёмки...")
        self.mw.progress.start()
        self.mw.master.update()
        
        def task():
            try:
                data, stats = process_survey_folder(
                    self.mw.survey_path.get(),
                    self.mw.mode.get(),
                    progress_callback=lambda msg: self.mw.master.after(0, self.mw.status_var.set, msg)
                )
                self.mw.survey_data = data
                self.mw.errors = stats.get('errors', [])
                self.mw.master.after(0, self._update_survey_preview, data, stats)
            except Exception as e:
                self.mw.master.after(0, messagebox.showerror, "Ошибка", str(e))
            finally:
                self.mw.master.after(0, self.mw.progress.stop)
                self.mw.master.after(0, self.mw.status_var.set, "Готов")
        
        threading.Thread(target=task, daemon=True).start()
    
    def _update_survey_preview(self, data, stats):
        self.mw.survey_table.clear()
        
        # Группируем листы по дате (первые 6 символов имени папки)
        date_groups = {}
        total_rows = 0
        for sheet_name, df in data.items():
            rows = len(df)
            total_rows += rows
            # Извлекаем дату из первых 6 символов, если возможно
            if len(sheet_name) >= 6:
                date_prefix = sheet_name[:6]   # YYMMDD
                date_key = f"20{date_prefix[0:2]}-{date_prefix[2:4]}-{date_prefix[4:6]}"
            else:
                date_key = sheet_name
            date_groups.setdefault(date_key, 0)
            date_groups[date_key] += rows

        # Заполняем таблицу: "День" – дата, "Строк" – сумма строк за день
        for date_key, sum_rows in sorted(date_groups.items()):
            self.mw.survey_table.insert_row([date_key, sum_rows])

        # Сохранение исходников (единый Excel)
        try:
            source_folder = Path(self.mw.output_dir.get()) / "Исходники"
            save_survey_excels(self.mw.survey_data, str(source_folder), self.mw.mode.get(),
                            nav_data=None, keep_only_matched=False)
        except Exception as e:
            self.mw.errors.append(f"Не удалось сохранить исходные файлы: {e}")

        # Обновление мини-карты съёмки
        if self.mw.survey_map_figure is not None:
            self.mw.survey_map_figure.clear()
            ax = self.mw.survey_map_figure.add_subplot(111)
            MapManager.draw_survey_track(ax, self.mw.survey_data)
            self.mw.survey_map_canvas.draw()

        msg = f"Съёмка загружена. Дней: {len(date_groups)}, всего строк: {total_rows}"
        if stats['errors']:
            msg += f"\nОшибок: {len(stats['errors'])} (нажмите 'Показать ошибки')"
        self.mw._add_statistics(msg)
        messagebox.showinfo("Готово", msg)
    
    def load_navigation(self):
        if not self.mw.nav_path.get():
            return
        self.mw.status_var.set("Загрузка навигационных данных...")
        self.mw.progress.start()
        self.mw.master.update()
        
        def task():
            try:
                data = process_navigation_folder(self.mw.nav_path.get())
                self.mw.nav_data = data
                self.mw.master.after(0, self._update_nav_preview, data)
            except Exception as e:
                self.mw.master.after(0, messagebox.showerror, "Ошибка", str(e))
            finally:
                self.mw.master.after(0, self.mw.progress.stop)
                self.mw.master.after(0, self.mw.status_var.set, "Готов")
        
        threading.Thread(target=task, daemon=True).start()
    
    def _update_nav_preview(self, data):
        self.mw.nav_table.clear()
        total_lines = 0
        for date_str, text in data.items():
            lines = text.count('\n') + 1 if text else 0
            total_lines += lines
            # Форматируем YYYYMMDD → YYYY-MM-DD
            if len(date_str) == 8:
                formatted_date = f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"
            else:
                formatted_date = date_str
            self.mw.nav_table.insert_row([formatted_date, lines])
        
        # Сохраняем объединённые навигационные файлы в Исходники
        if self.mw.output_dir.get():
            sources_folder = Path(self.mw.output_dir.get()) / "Исходники"
            sources_folder.mkdir(parents=True, exist_ok=True)
            for date_str, content in data.items():
                if len(date_str) == 8:
                    day = date_str[6:8]
                    month = date_str[4:6]
                    year = date_str[2:4]
                    short_date = f"{day}{month}{year}"
                else:
                    short_date = date_str
                suffix = "V1" if self.mw.mode.get() == "with_v1" else ""
                nav_file = sources_folder / f"{short_date}{suffix}.txt"
                try:
                    nav_file.write_text(content, encoding='utf-8')
                except Exception as e:
                    self.mw.errors.append(f"Не удалось сохранить навигацию {short_date}{suffix}: {e}")
        
        # Строим кэш координат для быстрой отрисовки карты
        self.mw.nav_coords_cache = {}
        for date_str, text in data.items():
            try:
                coords = parse_navigation_text(text)
                points = list(coords.values())
                self.mw.nav_coords_cache[date_str] = points
            except:
                self.mw.nav_coords_cache[date_str] = []
        
        if self.mw.nav_map_figure is not None:
            self.mw.nav_map_figure.clear()
            ax = self.mw.nav_map_figure.add_subplot(111)
            from gui.map_manager import MapManager
            MapManager.draw_nav_track(ax, self.mw.nav_coords_cache)
            self.mw.nav_map_canvas.draw()
        
        msg = f"Навигация загружена. Дат: {len(data)}, всего строк: {total_lines}"
        self.mw._add_statistics(msg)
        messagebox.showinfo("Готово", msg)
    
    def load_correction_preview(self, file_path):
        self.mw.status_var.set("Чтение файла вариаций...")
        self.mw.progress.start()
        self.mw.master.update()
        
        def task():
            try:
                import openpyxl
                # Открываем книгу в режиме только для чтения
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                sheet_stats = []
                total_rows = 0
                for sheet_name in sheets:
                    ws = wb[sheet_name]
                    rows = ws.max_row
                    # Если в листе есть заголовок, вычитаем 1 строку
                    if rows > 0:
                        rows -= 1
                    sheet_stats.append((sheet_name, rows))
                    total_rows += rows
                wb.close()
                self.mw.master.after(0, self._update_correction_preview, file_path, sheet_stats, len(sheets), total_rows)
            except Exception as e:
                self.mw.master.after(0, messagebox.showerror, "Ошибка", f"Не удалось прочитать файл вариаций:\n{e}")
            finally:
                self.mw.master.after(0, self.mw.progress.stop)
                self.mw.master.after(0, self.mw.status_var.set, "Готов")
        
        threading.Thread(target=task, daemon=True).start()
    
    def _update_correction_preview(self, file_path, sheet_stats, sheets_count, total_rows):
        self.mw.corr_table.clear()
        for sheet, rows in sheet_stats:
            self.mw.corr_table.insert_row([sheet, rows])
        self.mw.corr_label_var.set(f"Файл: {Path(file_path).name}")
        msg = f"Вариации загружены. Листов: {sheets_count}, всего строк: {total_rows}"
        self.mw._add_statistics(msg)
        messagebox.showinfo("Готово", msg)